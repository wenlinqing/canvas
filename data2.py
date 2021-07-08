# -*- coding: utf-8 -*-
import time

from pyquery import PyQuery as py
import requests
from requests.exceptions import RequestException
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.141 Safari/537.36'
}
# fileUrl = r'C:\Users\Administrator\Desktop\tkinter\dist\节点信息日报.xlsx'
fileUrl = r'C:\Users\Administrator\Desktop\dist\节点信息日报.xlsx'
# fileUrl = r'123.xlsx'

# indexUrl = r'C:\Users\Administrator\Desktop\tkinter\dist\index.html'
# indexUrl = r'index2.html'


def loading(obj):
    # url1 = 'https://filfox.info/zh/address/f097658'
    # url2 = 'https://filfox.info/zh/address/f0135885'
    url = 'https://filfox.info/zh/address/' + obj
    try:
        response = requests.get(url=url, headers=headers)
        if response.status_code == 200:
            return response.text
        else:
            return None
    except RequestException:
        print(RequestException)


def formatData(obj, objs):
    try:
        print(objs + ' 数据爬取中。。。')
        html = loading(obj)
        # print(html)
        doc = py(html)
        data1 = doc(
            '#__layout > div > div.container.mx-auto.flex-grow > div.hidden.lg\:block > div > div:nth-child(3) > div.grid.grid-rows-1.grid-cols-2.gap-4.my-4.mx-8 > div:nth-child(2) > div > div:nth-child(2) > p.font-medium.text-2xl').text()
        data2 = doc(
            '#__layout > div > div.container.mx-auto.flex-grow > div.hidden.lg\:block > div > div:nth-child(3) > div.grid.grid-rows-1.grid-cols-2.gap-4.my-4.mx-8 > div:nth-child(1) > div > div.py-4 > p.font-medium.text-2xl').text()
        data3 = doc(
            '#__layout > div > div.container.mx-auto.flex-grow > div.hidden.lg\:block > div > div:nth-child(3) > div.grid.grid-rows-1.grid-cols-2.gap-4.my-4.mx-8 > div:nth-child(1) > div > div.py-4 > p.flex.items-center.text-sm.mt-4').text()
        data4 = doc(
            '#__layout > div > div.container.mx-auto.flex-grow > div.hidden.lg\:block > div > div:nth-child(3) > div.grid.grid-rows-1.grid-cols-2.gap-4.my-4.mx-8 > div:nth-child(1) > div > div.py-4 > p:nth-child(4)').text()
        data5 = doc(
            '#__layout > div > div.container.mx-auto.flex-grow > div.hidden.lg\:block > div > div:nth-child(3) > div.grid.grid-rows-1.grid-cols-2.gap-4.my-4.mx-8 > div:nth-child(1) > div > div.py-4 > p:nth-child(5)').text()
        data6 = doc(
            '#__layout > div > div.container.mx-auto.flex-grow > div.hidden.lg\:block > div > div.bg-white.rounded-md.pb-6.my-4 > div.mx-8.border.border-background.rounded-sm.p-4 > div:nth-child(1) > div').text()
        data7 = doc(
            '#__layout > div > div.container.mx-auto.flex-grow > div.hidden.lg\:block > div > div.bg-white.rounded-md.pb-6.my-4 > div.mx-8.border.border-background.rounded-sm.p-4 > div:nth-child(1) > p:nth-child(2)').text()
        data8 = doc(
            '#__layout > div > div.container.mx-auto.flex-grow > div.hidden.lg\:block > div > div.bg-white.rounded-md.pb-6.my-4 > div.mx-8.border.border-background.rounded-sm.p-4 > div:nth-child(2) > div.text-sm.w-1\/4.text-left.flex.flex-row.items-center').text()
        data9 = doc(
            '#__layout > div > div.container.mx-auto.flex-grow > div.hidden.lg\:block > div > div.bg-white.rounded-md.pb-6.my-4 > div.mx-8.border.border-background.rounded-sm.p-4 > div:nth-child(2) > p').text()

        data10 = doc(
            '#__layout > div > div.container.mx-auto.flex-grow > div.hidden.lg\:block > div > div:nth-child(3) > div.grid.grid-rows-1.grid-cols-2.gap-4.my-4.mx-8 > div:nth-child(2) > div > div.text-sm.mt-2.items-center.justify-between.flex > div > span.text-green-600').text().strip()

        data11 = doc(
            '#__layout > div > div.container.mx-auto.flex-grow > div.hidden.lg\:block > div > div:nth-child(3) > div.grid.grid-rows-1.grid-cols-2.gap-4.my-4.mx-8 > div:nth-child(2) > div > div.text-sm.mt-2.items-center.justify-between.flex > div > span.text-red-700').text().strip()

        #  data1
        pib = data1.split(' ')[1].lower()
        if pib == 'tib':
            res = float(data1.split(' ')[0]) / 1024
            res = str(res)[:5]
        else:
            res = data1.split(' ')[0]
        d1 = res

        # data2
        d2 = data2.split(' ')[0]

        # data3
        d3 = data3.split(':')[1].split(' ')[1]

        # data4
        d4 = data4.split(' ')[1]

        # data5
        d5 = data5.split(' ')[1]

        # data6
        d6 = data6.split(':')[1].strip()

        # data7
        dd7_cell = data7.split(' ')[2].lower()
        dt = data7.split(' ')[1]
        if dd7_cell == 'pib':
            print('算力增速出现pib')
            d7 = float(dt) * 1024
        elif dd7_cell == 'gib':
            print('算力增速出现gib')
            d7 = float(dt) / 1024
        else:
            d7 = dt

        # data8
        d8 = data8.split(':')[1].strip()

        # data9
        d9 = data9.split(':')[1].strip().split(' ')[0]

        d10 = data10.split(' ')[0].replace(',', '')
        d11 = data11.split(' ')[0].replace(',', '')

        # cell1 = datetime.today().__format__('%Y年%m月%d日')
        d12 = datetime.today().__format__('%H:%M')

        data = [float(d1),
                float(d2.replace(',', '')),
                float(d3.replace(',', '')),
                float(d4.replace(',', '')),
                float(d5.replace(',', '')),
                float(d6),
                float(d7),
                float(d8),
                float(d9),
                int(d10),
                int(d11),
                d12]

        return data
    except Exception as e:
        print(e)


def writeFun(data):
    wb = load_workbook(fileUrl)
    ws = wb.active
    max_row = ws.max_row

    dates = datetime.today().__format__('%Y年%m月%d日')
    f_title = [dates, '节点号', '节点名称', '有效算力（P）', '账户余额（FIL）', '可用余额（FIL）', '扇区抵押（FIL）', '挖矿锁仓（FIL）', '矿机当量',
               '算力增速（T/24h）', '出块份数', '出块奖励（FIL）', '有效', '错误', '统计时间', '备注']

    mins = max_row + 2
    maxs = max_row + 2 + len(f_no)

    fill = PatternFill('solid', fgColor='d4d4d4')
    fill2 = PatternFill('solid', fgColor='FF9999')
    border = Border(
        left=Side(style='thin', color='FF000000'),
        right=Side(style='thin', color='FF000000'),
        top=Side(style='thin', color='FF000000'),
        bottom=Side(style='thin', color='FF000000'),
        diagonal=Side(style='thin', color='FF000000'),
        diagonal_direction=0, outline=Side(style='medium', color='FF000000'),
        vertical=Side(style='thin', color='FF000000'),
        horizontal=Side(style='thin', color='FF000000')
    )

    ws.append(f_title)
    ws.row_dimensions[mins].height = 30

    for b in range(2, 17):
        ws.cell(row=mins - 1, column=b).fill = fill
        ws.cell(row=mins - 1, column=b).alignment = Alignment(horizontal='center', vertical='center')
        # ws.cell(row=mins-1, column=b).border = border

    for i in range(0, len(data)):
        ws.append(data[i])

    fontBold = Font(bold=True)
    total = ['', '合计', '', '=SUM(D{}:D{})'.format(mins + 1, maxs), '=SUM(E{}:E{})'.format(mins + 1, maxs),
             '=SUM(F{}:F{})'.format(mins + 1, maxs),
             '=SUM(G{}:G{})'.format(mins + 1, maxs), '=SUM(H{}:H{})'.format(mins + 1, maxs),
             '=SUM(I{}:I{})'.format(mins + 1, maxs),
             '=SUM(J{}:J{})'.format(mins + 1, maxs), '=SUM(K{}:K{})'.format(mins + 1, maxs),
             '=SUM(L{}:L{})'.format(mins + 1, maxs), '', '', '']
    ws.append(total)
    ws.row_dimensions[maxs + 1].height = 25

    for x in range(mins - 1, maxs + 1):
        for y in range(2, 17):
            ws.cell(row=x, column=y).border = border
            ws.cell(row=x, column=y).alignment = Alignment(horizontal='center', vertical='center')
            if x > mins - 1 and y == 14:
                err = ws.cell(row=x, column=14).value
                # print(err, type(err))
                if err != '' and int(err) >= 1:
                    for b in range(2, 15):
                        ws.cell(row=x, column=b).fill = fill2

            if x == maxs:
                ws.cell(row=x, column=y).font = fontBold

    ws.merge_cells('A{}:A{}'.format(mins, maxs + 1))
    ws['A{}'.format(max_row + 1)].alignment = Alignment(horizontal='center', vertical='center')
    ws.insert_rows(max_row + 1)
    wb.save(fileUrl)  # 保存xlsx 保存xlsx 保存xlsx 保存xlsx 保存xlsx 保存xlsx 保存xlsx 保存xlsx 保存xlsx 保存xlsx 保存xlsx 保存xlsx

    [total1, total2, total3, total4, total5, total6, total7, total8, total9] = [0, 0, 0, 0, 0, 0, 0, 0, 0]
    for objs in data:
        # print(type(objs[3]),objs[3])  # 有效算力
        total1 += objs[3]
        total2 += objs[4]
        total3 += objs[5]
        total4 += objs[6]
        total5 += objs[7]
        total6 += objs[8]
        total7 += objs[9]
        total8 += objs[10]
        total9 += objs[11]
    total = ['', '合计', '',
             format(total1, '.3f'),
             format(total2, '.4f'),
             format(total3, '.4f'),
             format(total4, '.4f'),
             format(total5, '.4f'),
             format(total6, '.2f'),
             format(total7, '.4f'),
             format(total8, '.1f'),
             format(total9, '.4f'),
             '', '', '']

    # makeHtml(dates, f_title, data, total)  # 生成html 生成html 生成html 生成html 生成html 生成html 生成html 生成html 生成html 生成html
    print()
    print('数据爬取成功, 3秒后自动关闭')
    time.sleep(3)


#  以下生成html

strcss = '''
    <style>
      *{box-sizing: border-box;padding: 0;margin: 0;}
      body,html{padding: 0;margin: 0;font-size: 14px;}

      #app{padding:100px 0 0;margin: 0 auto;}
      .flex{display: flex;}
      .flex1{flex: 1}
      .leftBox{align-items: center;justify-content: center;min-width: 110px;font-size: 12px;}
      .thead{background: #f5f5f5;text-align: center;height: 30px;}
      table{border-collapse: collapse;width: 100%}
      td,th{border:1px solid #000;text-align: center;min-width: 100px;}

      .hasError{background: #ffb6b6}
      .fontBold{font-weight: bold;height:30px}
    </style>
'''


def makeHtml(time, f_title, dt, total):
    strhtml = """
        <!DOCTYPE HTML>
        <html>
           <head>
           <meta charset="utf-8">
           <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=0,user-scalable=no,minimal-ui">
            <title></title>
            {0}
           </head>
           <body>
              <div id="app">
                 <div class="Boxer flex">
                    <div class="flex leftBox">
                       {1}
                    </div>
                    <div class="flex1 rightBox">
                       <table cellspacing="0" cellpadding="0">{2}</table>
                    </div>
                 </div>
              </div>

           </body>
        </html>
    """.format(strcss, time, makeData(f_title, dt, total))

    # with open(indexUrl, 'w', encoding='utf-8') as f:
    #     f.write(strhtml)


def makeData(f_title, data, total):
    html = tableTh(f_title)
    for objs in data:
        if objs[len(objs) - 2] >= 1:
            html += tableTr('hasError', objs)
        else:
            html += tableTr('', objs)

    html += tableTr('fontBold', total)
    # print(html)
    return html


def tableTr(name, obj):
    str = r'<tr class="{}">'.format(name)
    for o in range(1, len(obj)):
        str += r'<td>{}</td>'.format(obj[o])
    str += '<td></td></tr>'

    return str


def tableTh(obj):
    str = r'<tr class="thead">'
    for o in range(1, len(obj)):
        str += r'<th>{}</th>'.format(obj[o])
    str += '</tr>'
    return str


# def tableTd(obj):
#     str = ''
#     for o in range(1, len(obj)):
#         str += r'<td>{}</td>'.format(obj[o])
#     return str + r'<td></td>'


if __name__ == '__main__':
    f_no = ['f097658', 'f0133833', 'f0135066', 'f0135885', 'f0136668', 'f0133689', 'f0135551', 'f0136138', 'f0843558',
            'f0842266', 'f0881281', 'f0881262', 'f0881531', 'f01032055', 'f01038625', 'f0881321']
    f_name = ['达客', '楠科', 'WELL-3', 'WELL-6', '金蟾国际～03', 'D&W-飞猪', '米粒矿场', '', 'D&W~旺旺狗', 'D&W～紫牛', '', '', '', '', '',
              '']

    result = []
    for i in range(0, len(f_no)):
        subDate = ['', f_no[i], f_name[i]]
        subDate.extend(formatData(f_no[i], f_no[i]))
        result.append(subDate)

    # print(result)

    writeFun(result)
