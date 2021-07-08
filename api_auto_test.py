import json
import time

import requests
from openpyxl import load_workbook, Workbook


class autoTest():
    def __init__(self):
        self.api_list = []
        self.get_xls_data()
        self.test_ok = 0
        self.test_error = 0
        self.test_error_ids = []
        self.token = None

    def get_xls_data(self):
        try:
            wb = load_workbook('api_auto_test.xlsx')
        except Exception as e:
            print('读取测试用例文件失败 ===', e)
        else:
            sheet = wb.active
            for n in range(2, sheet.max_row):
                value = sheet.cell(n, 7).value  # 是否需要测试
                if value == 1:
                    # print(value, sheet.cell(n, 8).value)
                    self.api_list.append({
                        'id': sheet.cell(n, 1).value,
                        'url': sheet.cell(n, 5).value,
                        'data': sheet.cell(n, 9).value,
                        'hope_result': sheet.cell(n, 10).value
                    })
                else:
                    continue
            # print(self.api_list)

    def api_fun(self):
        try:
            print('开始接口测试')
            for index in range(0, len(self.api_list)):
                obj = self.api_list[index]
                headers = {}
                if self.token:
                    headers['Authorization'] = 'Bearer ' + self.token

                backs = requests.post(url=obj['url'], data=json.loads(obj['data']), headers=headers)
                # print('results.request.headers', backs.headers)
                results = backs.json()
                hope_result = eval(obj['hope_result'])
                # print('results===', results)
                if results['code'] == hope_result['code'] and results['msg'] == hope_result['msg']:
                    print('ID为 {0} api 测试通过'.format(obj['id']))
                    text = 'api test pass'
                    self.test_ok += 1
                    if 'data' in results:
                        # print('data==', results['data']['token'])
                        self.token = results['data']['token']
                else:
                    print('ID为 {0} api 测试失败 测试失败 测试失败'.format(obj['id']))
                    text = 'api test errors'
                    self.test_error += 1
                    self.test_error_ids.append(obj['id'])  # 记录测试失败的id
                # time.sleep(1)

            print('测试结束 总共测试接口：{0} 条， 成功{1}条，失败{2}条 {3}'
                  .format(self.test_ok + self.test_error, self.test_ok, self.test_error, self.test_error_ids))
            self.write_to_file()
        except Exception as e:
            print('eeeeeeeeeeeee=', e)

    def write_to_file(self):
        # wb = Workbook()
        # sheet = wb.create_sheet('测试结果')
        # sheet['A1'] = '总共测试接口：{0} 条， 成功{1}条，失败{2}条 {3}'.format(self.test_ok + self.test_error, self.test_ok,
        #                                                        self.test_error, self.test_error_ids)
        # sheet['A2'] = '测试日期: {}'.format(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
        # # print("sheet['A1'] = ", sheet['A1'].value )
        # wb.save('api_auto_test2.xlsx')
        try:
            with open('api_auto_test.txt', 'w', encoding='utf-8') as f:
                f.write('\n总共测试接口：{0} 条， 成功{1}条，失败{2}条 {3} \n\n'.format(self.test_ok + self.test_error, self.test_ok,
                                                                        self.test_error, self.test_error_ids))
                f.write('测试日期: {}'.format(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())))
            print('已生成测试报告')
        except Exception as e:
            print('生成测试报告错误==', e)


test = autoTest()
test.api_fun()
# test.write_to_file()
